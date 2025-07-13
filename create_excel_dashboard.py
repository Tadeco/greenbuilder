import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart

# Cria o workbook e as abas
wb = Workbook()

# Aba 1: Resumo Nacional
data_resumo = [
    ["Indicador", "Valor"],
    ["Área Territorial (km²)", "8.510.296"],
    ["População Estimada", "210.147.125"],
    ["PIB Total (R$ 1.000)", "6.093.497,48"],
    ["PIB per Capita (R$)", "28.996,34"],
    ["Unidades da Federação", "27"],
    ["Municípios", "5.570"],
]
ws1 = wb.active
ws1.title = "Resumo Nacional"
for row in data_resumo:
    ws1.append(row)

# Aba 2: Situação das Declarações
data_situacao = [
    ["Macrorregião", "Estados Declarantes", "% Estados", "Municípios Declarantes", "% Municípios"],
    ["Norte", 7, 100, 248, 55.11],
    ["Nordeste", 8, 88.89, 651, 36.29],
    ["Sudeste", 4, 100, 737, 44.18],
    ["Sul", 3, 100, 600, 50.38],
    ["Centro-Oeste", 4, 100, 256, 54.82],
]
ws2 = wb.create_sheet("Situação Declarações")
for row in data_situacao:
    ws2.append(row)

# Aba 3: Planos de Gestão
data_planos = [
    ["Indicador", "Valor"],
    ["Planos estaduais segundo a PNRS", "20 (74,07%)"],
    ["Planos municipais segundo a PNRS", "-"],
    ["Planos intermunicipais segundo a PNRS", "-"],
]
ws3 = wb.create_sheet("Planos de Gestão")
for row in data_planos:
    ws3.append(row)

# Aba 4: Gestão Compartilhada
data_gestao = [
    ["Estado", "% Municípios em Soluções Compartilhadas"],
    ["(dados não disponíveis no sumário)", ""],
]
ws4 = wb.create_sheet("Gestão Compartilhada")
for row in data_gestao:
    ws4.append(row)

# Aba 5: Autossuficiência Financeira
data_auto = [
    ["Macrorregião", "Custo Total por Habitante (R$/hab)"],
    ["Norte", 0],
    ["Nordeste", 0],
    ["Sudeste", 0],
    ["Sul", 0],
    ["Centro-Oeste", 0],
]
ws5 = wb.create_sheet("Autossuficiência Fin.")
for row in data_auto:
    ws5.append(row)

# Aba 6: Resíduos Sólidos Urbanos
data_residuos = [
    ["Indicador", "Valor"],
    ["Massa total coletada (t)", 0],
    ["Massa da coleta seletiva (t)", 0],
    ["Cobertura da coleta seletiva (%)", 0],
    [],
    ["Macrorregião", "Aterros Sanitários (t)", "Lixões/Aterros Controlados (t)", "Total (t)"],
    ["Norte", 0, 0, 0],
    ["Nordeste", 0, 0, 0],
    ["Sudeste", 0, 0, 0],
    ["Sul", 0, 0, 0],
    ["Centro-Oeste", 0, 0, 0],
]
ws6 = wb.create_sheet("Resíduos Urbanos")
for row in data_residuos:
    ws6.append(row)

# Dashboard com gráficos
ws_dash = wb.create_sheet("Dashboard")
ws_dash["A1"] = "Dashboard Interativa - Resíduos Sólidos 2020"

# Gráfico de Coluna: Municípios Declarantes por Macrorregião
bar_chart = BarChart()
bar_chart.title = "Municípios Declarantes por Macrorregião"
bar_chart.x_axis.title = "Macrorregião"
bar_chart.y_axis.title = "Municípios Declarantes"
data = Reference(ws2, min_col=4, min_row=1, max_row=6)
cats = Reference(ws2, min_col=1, min_row=2, max_row=6)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
ws_dash.add_chart(bar_chart, "A3")

# Gráfico de Barra: % Municípios Declarantes por Macrorregião
bar_chart2 = BarChart()
bar_chart2.type = "bar"
bar_chart2.title = "% Municípios Declarantes por Macrorregião"
bar_chart2.x_axis.title = "Macrorregião"
bar_chart2.y_axis.title = "% Municípios Declarantes"
data2 = Reference(ws2, min_col=5, min_row=1, max_row=6)
cats2 = Reference(ws2, min_col=1, min_row=2, max_row=6)
bar_chart2.add_data(data2, titles_from_data=True)
bar_chart2.set_categories(cats2)
ws_dash.add_chart(bar_chart2, "J3")

# Gráfico de Pizza: Distribuição dos Estados Declarantes
pie = PieChart()
pie.title = "Distribuição dos Estados Declarantes"
data3 = Reference(ws2, min_col=2, min_row=2, max_row=6)
labels = Reference(ws2, min_col=1, min_row=2, max_row=6)
pie.add_data(data3, titles_from_data=False)
pie.set_categories(labels)
ws_dash.add_chart(pie, "A20")

# Salva o arquivo
wb.save("senirgraficos.xlsx") 