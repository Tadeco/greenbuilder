import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart, LineChart

wb = Workbook()

abas = [
    "Resumo Nacional",
    "Situação das Declarações",
    "Planos de Gestão",
    "Gestão Compartilhada",
    "Autossuficiência Financeira",
    "Resíduos Sólidos Urbanos",
    "Resíduos de Serviços de Saúde",
    "Resíduos de Construção Civil",
    "Resíduos de Serviços Públicos de Saneamento Básico",
    "Resíduos de Serviços de Transporte",
    "Resíduos Industriais",
    "Resíduos de Mineração",
    "Outros Resíduos de Atividades Potencialmente Poluidoras",
    "Logística Reversa",
    "Resumo das Informações Estaduais"
]

cabecalho = [
    "Ano", "Estado", "Macrorregião", "Tipo de Resíduo", "Quantidade Gerada (t)", "Quantidade Destinada (t)", "Quantidade Reciclada (t)", "Municípios Participantes", "Observações"
]

dados_exemplo = {
    "Resumo Nacional": [
        [2019, "BR", "Nacional", "Todos", 500000, 480000, 150000, 5570, "Total Brasil"],
        [2019, "BR", "Nacional", "Todos", 510000, 490000, 160000, 5570, "Total Brasil"],
        [2019, "BR", "Nacional", "Todos", 520000, 500000, 170000, 5570, "Total Brasil"],
        [2019, "BR", "Nacional", "Todos", 530000, 510000, 180000, 5570, "Total Brasil"],
        [2019, "BR", "Nacional", "Todos", 540000, 520000, 190000, 5570, "Total Brasil"],
    ],
    "Situação das Declarações": [
        [2019, "SP", "Sudeste", "Declaração", 645, 600, 200, 645, ""],
        [2019, "RJ", "Sudeste", "Declaração", 92, 90, 30, 92, ""],
        [2019, "MG", "Sudeste", "Declaração", 853, 800, 250, 853, ""],
        [2019, "BA", "Nordeste", "Declaração", 417, 400, 120, 417, ""],
        [2019, "RS", "Sul", "Declaração", 497, 480, 150, 497, ""],
    ],
    "Planos de Gestão": [
        [2019, "SP", "Sudeste", "Plano Estadual", 1, 1, 1, 1, "Aprovado"],
        [2019, "RJ", "Sudeste", "Plano Municipal", 50, 45, 20, 50, "Em revisão"],
        [2019, "MG", "Sudeste", "Plano Intermunicipal", 10, 9, 5, 10, ""],
        [2019, "BA", "Nordeste", "Plano Estadual", 1, 1, 1, 1, "Aprovado"],
        [2019, "RS", "Sul", "Plano Municipal", 30, 28, 10, 30, ""],
    ],
    "Gestão Compartilhada": [
        [2019, "SP", "Sudeste", "Consórcio", 5, 5, 2, 100, ""],
        [2019, "RJ", "Sudeste", "Consórcio", 3, 3, 1, 60, ""],
        [2019, "MG", "Sudeste", "Consórcio", 4, 4, 2, 80, ""],
        [2019, "BA", "Nordeste", "Consórcio", 2, 2, 1, 40, ""],
        [2019, "RS", "Sul", "Consórcio", 3, 3, 1, 50, ""],
    ],
    "Autossuficiência Financeira": [
        [2019, "SP", "Sudeste", "Financeiro", 1000000, 950000, 300000, 645, ""],
        [2019, "RJ", "Sudeste", "Financeiro", 500000, 480000, 150000, 92, ""],
        [2019, "MG", "Sudeste", "Financeiro", 200000, 195000, 50000, 853, ""],
        [2019, "BA", "Nordeste", "Financeiro", 300000, 290000, 80000, 417, ""],
        [2019, "RS", "Sul", "Financeiro", 400000, 390000, 120000, 497, ""],
    ],
    "Resíduos Sólidos Urbanos": [
        [2019, "SP", "Sudeste", "RSU", 100000, 95000, 30000, 645, ""],
        [2019, "RJ", "Sudeste", "RSU", 50000, 48000, 15000, 92, ""],
        [2019, "MG", "Sudeste", "RSU", 20000, 19500, 5000, 853, ""],
        [2019, "BA", "Nordeste", "RSU", 30000, 29000, 8000, 417, ""],
        [2019, "RS", "Sul", "RSU", 40000, 39000, 12000, 497, ""],
    ],
    "Resíduos de Serviços de Saúde": [
        [2019, "SP", "Sudeste", "Saúde", 8000, 7900, 2000, 645, ""],
        [2019, "RJ", "Sudeste", "Saúde", 4000, 3900, 1000, 92, ""],
        [2019, "MG", "Sudeste", "Saúde", 6000, 5900, 1500, 853, ""],
        [2019, "BA", "Nordeste", "Saúde", 3000, 2900, 800, 417, ""],
        [2019, "RS", "Sul", "Saúde", 5000, 4900, 1200, 497, ""],
    ],
    "Resíduos de Construção Civil": [
        [2019, "SP", "Sudeste", "Construção", 20000, 19500, 5000, 645, ""],
        [2019, "RJ", "Sudeste", "Construção", 10000, 9800, 2500, 92, ""],
        [2019, "MG", "Sudeste", "Construção", 15000, 14800, 3500, 853, ""],
        [2019, "BA", "Nordeste", "Construção", 8000, 7900, 2000, 417, ""],
        [2019, "RS", "Sul", "Construção", 12000, 11900, 3000, 497, ""],
    ],
    "Resíduos de Serviços Públicos de Saneamento Básico": [
        [2019, "SP", "Sudeste", "Saneamento", 5000, 4900, 1200, 645, ""],
        [2019, "RJ", "Sudeste", "Saneamento", 2500, 2450, 600, 92, ""],
        [2019, "MG", "Sudeste", "Saneamento", 4000, 3950, 900, 853, ""],
        [2019, "BA", "Nordeste", "Saneamento", 2000, 1980, 400, 417, ""],
        [2019, "RS", "Sul", "Saneamento", 3000, 2980, 700, 497, ""],
    ],
    "Resíduos de Serviços de Transporte": [
        [2019, "SP", "Sudeste", "Transporte", 3000, 2950, 800, 645, ""],
        [2019, "RJ", "Sudeste", "Transporte", 1500, 1480, 400, 92, ""],
        [2019, "MG", "Sudeste", "Transporte", 2000, 1980, 500, 853, ""],
        [2019, "BA", "Nordeste", "Transporte", 1000, 990, 200, 417, ""],
        [2019, "RS", "Sul", "Transporte", 1800, 1780, 300, 497, ""],
    ],
    "Resíduos Industriais": [
        [2019, "SP", "Sudeste", "Industrial", 40000, 39000, 12000, 645, ""],
        [2019, "RJ", "Sudeste", "Industrial", 20000, 19800, 6000, 92, ""],
        [2019, "MG", "Sudeste", "Industrial", 30000, 29800, 9000, 853, ""],
        [2019, "BA", "Nordeste", "Industrial", 16000, 15900, 4000, 417, ""],
        [2019, "RS", "Sul", "Industrial", 24000, 23900, 7000, 497, ""],
    ],
    "Resíduos de Mineração": [
        [2019, "MG", "Sudeste", "Mineração", 100000, 99000, 20000, 853, ""],
        [2019, "PA", "Norte", "Mineração", 80000, 79000, 15000, 144, ""],
        [2019, "GO", "Centro-Oeste", "Mineração", 60000, 59000, 12000, 246, ""],
        [2019, "BA", "Nordeste", "Mineração", 40000, 39500, 8000, 417, ""],
        [2019, "RS", "Sul", "Mineração", 20000, 19800, 4000, 497, ""],
    ],
    "Outros Resíduos de Atividades Potencialmente Poluidoras": [
        [2019, "SP", "Sudeste", "Outros", 5000, 4900, 1200, 645, ""],
        [2019, "RJ", "Sudeste", "Outros", 2500, 2450, 600, 92, ""],
        [2019, "MG", "Sudeste", "Outros", 4000, 3950, 900, 853, ""],
        [2019, "BA", "Nordeste", "Outros", 2000, 1980, 400, 417, ""],
        [2019, "RS", "Sul", "Outros", 3000, 2980, 700, 497, ""],
    ],
    "Logística Reversa": [
        [2019, "SP", "Sudeste", "Logística Reversa", 1000, 950, 400, 645, ""],
        [2019, "RJ", "Sudeste", "Logística Reversa", 500, 480, 200, 92, ""],
        [2019, "MG", "Sudeste", "Logística Reversa", 800, 790, 300, 853, ""],
        [2019, "BA", "Nordeste", "Logística Reversa", 400, 390, 150, 417, ""],
        [2019, "RS", "Sul", "Logística Reversa", 600, 590, 250, 497, ""],
    ],
    "Resumo das Informações Estaduais": [
        [2019, "SP", "Sudeste", "Resumo", 200000, 195000, 60000, 645, ""],
        [2019, "RJ", "Sudeste", "Resumo", 100000, 98000, 30000, 92, ""],
        [2019, "MG", "Sudeste", "Resumo", 150000, 148000, 45000, 853, ""],
        [2019, "BA", "Nordeste", "Resumo", 80000, 79000, 20000, 417, ""],
        [2019, "RS", "Sul", "Resumo", 120000, 119000, 35000, 497, ""],
    ],
}

# Primeira aba ativa
ws = wb.active
ws.title = abas[0]
ws.append(cabecalho)
for row in dados_exemplo[abas[0]]:
    ws.append(row)

# Demais abas
sheets = {abas[0]: wb.active}
for aba in abas[1:]:
    wsx = wb.create_sheet(aba)
    wsx.append(cabecalho)
    for row in dados_exemplo[aba]:
        wsx.append(row)
    sheets[aba] = wsx

# Criar uma aba para gráficos
ws_graph = wb.create_sheet("Gráficos")
ws_graph["A1"] = "Gráficos de Sugestão para Power BI"

row_offset = 3
# Para armazenar taxas de reciclagem por aba
taxas_reciclagem = []

for aba in abas:
    ws_data = sheets[aba]
    # Gráfico de barras: Quantidade Gerada por Estado
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = f"{aba} - Quantidade Gerada por Estado"
    bar_chart.x_axis.title = "Estado"
    bar_chart.y_axis.title = "Quantidade Gerada (t)"
    data = Reference(ws_data, min_col=5, min_row=1, max_row=6)
    cats = Reference(ws_data, min_col=2, min_row=2, max_row=6)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    ws_graph.add_chart(bar_chart, f"A{row_offset}")
    row_offset += 15
    # Gráfico de pizza: Proporção de resíduos por tipo
    pie = PieChart()
    pie.title = f"{aba} - Proporção por Tipo"
    data_pie = Reference(ws_data, min_col=5, min_row=2, max_row=6)
    labels_pie = Reference(ws_data, min_col=4, min_row=2, max_row=6)
    pie.add_data(data_pie, titles_from_data=False)
    pie.set_categories(labels_pie)
    ws_graph.add_chart(pie, f"J{row_offset-15}")
    # Gráfico de linha: Evolução anual da geração de resíduos
    line_chart = LineChart()
    line_chart.title = f"{aba} - Evolução Anual"
    line_chart.x_axis.title = "Ano"
    line_chart.y_axis.title = "Quantidade Gerada (t)"
    data_line = Reference(ws_data, min_col=5, min_row=1, max_row=6)
    cats_line = Reference(ws_data, min_col=1, min_row=2, max_row=6)
    line_chart.add_data(data_line, titles_from_data=True)
    line_chart.set_categories(cats_line)
    ws_graph.add_chart(line_chart, f"A{row_offset+8}")
    row_offset += 10
    # Calcular taxa de reciclagem da aba
    total_gerado = sum([row[4] for row in dados_exemplo[aba]])
    total_reciclado = sum([row[6] for row in dados_exemplo[aba]])
    taxa = total_reciclado / total_gerado if total_gerado else 0
    taxas_reciclagem.append((aba, taxa))

# Gráfico de barras para Taxa de Reciclagem por aba
ws_graph[f"A{row_offset+10}"] = "Taxa de Reciclagem por Aba"
for idx, (aba, taxa) in enumerate(taxas_reciclagem):
    ws_graph[f"A{row_offset+11+idx}"] = aba
    ws_graph[f"B{row_offset+11+idx}"] = taxa
bar_chart_taxa = BarChart()
bar_chart_taxa.type = "col"
bar_chart_taxa.title = "Taxa de Reciclagem por Aba"
bar_chart_taxa.x_axis.title = "Aba"
bar_chart_taxa.y_axis.title = "Taxa de Reciclagem"
data_taxa = Reference(ws_graph, min_col=2, min_row=row_offset+11, max_row=row_offset+11+len(taxas_reciclagem)-1)
cats_taxa = Reference(ws_graph, min_col=1, min_row=row_offset+11, max_row=row_offset+11+len(taxas_reciclagem)-1)
bar_chart_taxa.add_data(data_taxa, titles_from_data=False)
bar_chart_taxa.set_categories(cats_taxa)
ws_graph.add_chart(bar_chart_taxa, f"D{row_offset+10}")

wb.save("SINIRATT2019.xlsx") 